from openpyxl import load_workbook
workbook = load_workbook(filename="Sem5.xlsx")
sheet = workbook.active

marks = []
for col in sheet['AG']:
    marks.append(col.value)

names = []
for col in sheet['AH']:
    names.append(col.value)

marks = [i for i in marks if i != None]
names = [i for i in names if i != None]

print(len(names))
print(len(marks))

#data = list(zip(names,marks))
# print(data)


filename = "Result.xlsx"
workbookr = load_workbook(filename=filename)

sheetr = workbookr.active

name = input('Enter your full name(lastname firstname fathersname mothersname) in ALL CAPS (females add an / before your name e.g /RIYA ... )\n')

if name in names:
    ind = names.index(name)
    print(ind)
    num = marks[ind]
    print(num)
else:
    print('error! check again')

sheetr["A4"] = num
workbookr.save(filename=filename)

